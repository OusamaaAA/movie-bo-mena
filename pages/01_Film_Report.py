"""Film Performance Report — week-by-week box office with charts per market."""
from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from datetime import datetime

import altair as alt
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as PDFImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select

from src.db import session_scope
from src.models import FilmInvestmentAnalysis, FilmPerformanceFeatures, RatingsMetric
from src.repositories.film_repository import FilmRepository
from src.repositories.report_repository import ReportRepository
from src.services.period_display import dedupe_bom_weekend_rows, format_period, format_period_row
from src.services.inference import (
    InferenceInput,
    predict_via_api,
    run_investment_analysis,
    suggest_spend_from_target,
)
from src.services.ratings_display import format_ratings_line
from src.services.report_builder import build_film_report
from src.services.semantics import period_key_sort_ordinal, period_key_to_iso_week
from src.services.ticket_pricing import (
    estimate_admissions,
    load_prices,
    price_basis_label,
    sum_gross_in_ticket_currency,
    ticket_currency_code,
)
from src.config import get_settings
from ui_helpers import (
    inject_global_css,
    market_name,
    format_gross,
    page_header,
    section_divider,
    section_header,
    kpi_card,
    decision_badge,
    empty_state,
)

inject_global_css()

# ── helpers ──────────────────────────────────────────────────────────────────

MARKET_NAMES = {
    "EG": "Egypt", "SA": "Saudi Arabia", "AE": "UAE",
    "KW": "Kuwait", "BH": "Bahrain", "QA": "Qatar",
    "OM": "Oman", "JO": "Jordan", "LB": "Lebanon",
}


def source_label(source: str, granularity: str) -> str:
    g = (granularity or "").lower()
    s = source or "-"
    if g in ("day", "daily"):
        return f"{s} (daily)"
    if g in ("week", "weekly"):
        return f"{s} (weekly)"
    if g == "weekend":
        return f"{s} (weekend)"
    return s


def _market_totals(rows: list[dict]) -> list[dict]:
    """Per-market totals in **ticket baseline** currency (BOM USD rows converted per market)."""
    by_code: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        code = r.get("country_code") or ""
        if code:
            by_code[code].append(r)
    totals: dict[str, dict] = {}
    for code, mrows in by_code.items():
        tcur = ticket_currency_code(code) or ""
        conv_total = sum_gross_in_ticket_currency(mrows, code)
        latest_pk = ""
        periods = len(mrows)
        for r in mrows:
            pk = str(r.get("period_key") or "")
            if period_key_sort_ordinal(pk) > period_key_sort_ordinal(latest_pk):
                latest_pk = pk
        totals[code] = {
            "total": conv_total,
            "currency": tcur,
            "periods": periods,
            "latest_pk": latest_pk,
        }
    return [
        {
            "Market": market_name(code),
            "Total Gross": format_gross(t["total"], t["currency"]),
            "Periods": t["periods"],
            "Latest Period": format_period(t["latest_pk"]),
        }
        for code, t in sorted(totals.items(), key=lambda x: x[1]["total"], reverse=True)
    ]


def _row_value_for_chart(
    r: dict,
    country_code: str,
    metric: str,
    egypt_ticket_price: float | None,
) -> float | None:
    """Return gross or admissions for a reconciled row; admissions may be estimated from gross."""
    code = (country_code or "").upper()
    gross = float(r.get("period_gross_local") or 0)
    if metric == "gross":
        return gross if gross > 0 else None
    adm = r.get("admissions_actual")
    if adm is not None and float(adm) > 0:
        return float(adm)
    if gross <= 0:
        return None
    price_ov = egypt_ticket_price if code == "EG" else None
    gcur = (r.get("currency") or "").strip() or None
    est = estimate_admissions(gross, code, price_ov, gross_currency=gcur)
    return float(est) if est else None


def _chart_df(
    rows: list[dict],
    country_code: str,
    metric: str,
    egypt_ticket_price: float | None,
) -> pd.DataFrame | None:
    filtered = [r for r in rows if (r.get("country_code") or "").upper() == country_code.upper()]
    if not filtered:
        return None
    records = []
    for r in filtered:
        val = _row_value_for_chart(r, country_code, metric, egypt_ticket_price)
        if val is None or val <= 0:
            continue
        records.append(
            {
                "label": format_period_row(r),
                "ordinal": period_key_sort_ordinal(str(r.get("period_key") or "")),
                "value": val,
            }
        )
    if not records:
        return None
    return pd.DataFrame(records).sort_values("ordinal").drop_duplicates("label")


def _sorted_label_chart(
    data: pd.DataFrame,
    value_col: str,
    *,
    y_title: str = "Gross",
    color_col: str | None = None,
    height: int = 220,
) -> alt.Chart:
    base = alt.Chart(data)
    if color_col:
        return (
            base.mark_line(point=True)
            .encode(
                x=alt.X("label:N", sort=alt.EncodingSortField(field="ordinal", order="ascending"), title="Period"),
                y=alt.Y(f"{value_col}:Q", title=y_title),
                color=alt.Color(f"{color_col}:N", title="Source"),
                tooltip=["label:N", f"{value_col}:Q", f"{color_col}:N"],
            )
            .properties(height=height)
        )
    return (
        base.mark_line(point=True)
        .encode(
            x=alt.X("label:N", sort=alt.EncodingSortField(field="ordinal", order="ascending"), title="Period"),
            y=alt.Y(f"{value_col}:Q", title=y_title),
            tooltip=["label:N", f"{value_col}:Q"],
        )
        .properties(height=height)
    )


def _aggregate_filmyard_daily_to_weekly(filmyard_rows: list[dict]) -> list[dict]:
    """Aggregate Filmyard DAILY rows into ISO-week buckets."""
    weekly: dict[str, dict] = {}
    for r in filmyard_rows:
        if (r.get("record_granularity") or "").lower() not in ("day", "daily"):
            continue
        pk = str(r.get("period_key") or "")
        iso_week = period_key_to_iso_week(pk)
        if not iso_week:
            continue
        gross = float(r.get("period_gross_local") or 0)
        adm = r.get("admissions_actual")
        curr = r.get("currency") or ""
        if iso_week not in weekly:
            weekly[iso_week] = {"gross": 0.0, "admissions": 0.0, "has_admissions": False, "currency": curr}
        weekly[iso_week]["gross"] += gross
        if adm is not None:
            weekly[iso_week]["admissions"] += float(adm)
            weekly[iso_week]["has_admissions"] = True
    return [
        {
            "period_key": wk,
            "period_gross_local": v["gross"],
            "admissions_actual": v["admissions"] if v["has_admissions"] else None,
            "currency": v["currency"],
            "record_granularity": "week",
            "source": "Filmyard (daily agg.)",
        }
        for wk, v in sorted(weekly.items(), key=lambda x: period_key_sort_ordinal(x[0]))
        if v["gross"] > 0
    ]


def _calc_egypt_ticket_price(filmyard_rows: list[dict]) -> float | None:
    """Derive average ticket price from Filmyard rows with both gross and admissions."""
    prices = []
    for r in filmyard_rows:
        gross = float(r.get("period_gross_local") or 0)
        adm = r.get("admissions_actual")
        if adm is None or gross <= 0:
            continue
        adm_f = float(adm)
        if adm_f > 0:
            prices.append(gross / adm_f)
    if not prices:
        return None
    prices.sort()
    n = len(prices)
    mid = n // 2
    return prices[mid] if n % 2 == 1 else (prices[mid - 1] + prices[mid]) / 2


def _build_market_chart_png(title_text: str, labels: list[str], series: dict[str, list[float]], y_title: str) -> bytes | None:
    if not labels or not series:
        return None
    fig, ax = plt.subplots(figsize=(8.2, 3.4))
    for name, vals in series.items():
        if not vals:
            continue
        ax.plot(labels, vals, marker="o", linewidth=2, label=name)
    ax.set_title(title_text)
    ax.set_ylabel(y_title)
    ax.set_xlabel("Period")
    ax.grid(True, alpha=0.3)
    if len(series) > 1:
        ax.legend(loc="best")
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right")
    fig.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=160)
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _build_excel_report(
    film_title: str,
    summary_rows: list[dict],
    totals_rows: list[dict],
    detail_rows: list[dict],
    marketing_rows: list[dict],
    target_rows: list[dict],
    investment_rows: list[dict],
    target_result_rows: list[dict],
    chart_images: dict[str, bytes],
) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Film Performance Report"])
    ws_summary.append([])
    ws_summary.append(["Film", film_title])
    for row in summary_rows:
        ws_summary.append([row["Item"], row["Value"]])
    ws_summary["A1"].font = Font(bold=True, color="FFFFFFFF", size=13)
    ws_summary["A1"].fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    for col in ("A", "B"):
        ws_summary.column_dimensions[col].width = 38
    for cell in ws_summary[3]:
        cell.font = Font(bold=True)

    def _add_table_sheet(name: str, rows: list[dict]) -> None:
        if not rows:
            return
        ws = wb.create_sheet(name)
        headers = list(rows[0].keys())
        ws.append(headers)
        for hcell in ws[1]:
            hcell.font = Font(bold=True, color="FFFFFFFF")
            hcell.fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + idx)].width = 24

    _add_table_sheet("Market Totals", totals_rows)
    _add_table_sheet("Detailed Period Data", detail_rows)
    _add_table_sheet("Marketing Spend", marketing_rows)
    _add_table_sheet("Admissions Targets", target_rows)
    _add_table_sheet("Investment Analysis", investment_rows)
    _add_table_sheet("Target First Watch", target_result_rows)

    if chart_images:
        ws_charts = wb.create_sheet("Charts")
        ws_charts["A1"] = "Trend Charts"
        ws_charts["A1"].font = Font(bold=True, color="FFFFFFFF", size=12)
        ws_charts["A1"].fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
        row_pos = 3
        for chart_name, chart_png in chart_images.items():
            ws_charts[f"A{row_pos}"] = chart_name
            ws_charts[f"A{row_pos}"].font = Font(bold=True)
            img_buf = BytesIO(chart_png)
            ximg = XLImage(img_buf)
            ximg.width = 760
            ximg.height = 300
            ws_charts.add_image(ximg, f"A{row_pos + 1}")
            row_pos += 18

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def _build_pdf_report(
    film_title: str,
    summary_rows: list[dict],
    totals_rows: list[dict],
    investment_rows: list[dict],
    target_result_rows: list[dict],
    chart_images: dict[str, bytes],
) -> bytes:
    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, leftMargin=1.3 * cm, rightMargin=1.3 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>Film Performance Report</b>: {film_title}", styles["Title"]))
    story.append(Spacer(1, 0.25 * cm))

    sum_tbl = Table([["Item", "Value"]] + [[r["Item"], r["Value"]] for r in summary_rows], colWidths=[6.0 * cm, 11.5 * cm])
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(sum_tbl)

    def _append_table(title_text: str, rows: list[dict]) -> None:
        if not rows:
            return
        story.append(Spacer(1, 0.35 * cm))
        story.append(Paragraph(f"<b>{title_text}</b>", styles["Heading3"]))
        headers = list(rows[0].keys())
        tbl_data = [headers] + [[str(r.get(h, "")) for h in headers] for r in rows]
        col_w = 17.5 / max(1, len(headers))
        tbl = Table(tbl_data, colWidths=[col_w * cm] * len(headers))
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(tbl)

    _append_table("Box Office by Market", totals_rows)
    _append_table("Investment Analysis", investment_rows)
    _append_table("Target First Watch", target_result_rows)

    if chart_images:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph("<b>Trend Charts</b>", styles["Heading3"]))
        for chart_name, chart_png in chart_images.items():
            story.append(Paragraph(chart_name, styles["BodyText"]))
            story.append(Spacer(1, 0.1 * cm))
            story.append(PDFImage(BytesIO(chart_png), width=17.2 * cm, height=7.0 * cm))
            story.append(Spacer(1, 0.25 * cm))

    doc.build(story)
    out.seek(0)
    return out.getvalue()


# ── UI ───────────────────────────────────────────────────────────────────────

page_header("Film Performance", "Search, analyse, and run investment analysis on any tracked film.", icon="📊")

query = st.text_input("Search by title", placeholder="e.g. Siko Siko")
if not query:
    empty_state("Type a film title to begin.")
    st.stop()

with session_scope() as session:
    films = FilmRepository(session).search_rows(query, limit=10)

if not films:
    st.warning(
        "No films found in the database. "
        "Use **Live Data Fetch** (page 2 in the sidebar) to fetch this title from live sources."
    )
    st.stop()

labels = [f"{f['canonical_title']} ({f['release_year'] or '-'})" for f in films]
selected_label = st.selectbox("Select film", labels)
selected_film_id = films[labels.index(selected_label)]["id"]

with session_scope() as session:
    film = FilmRepository(session).get(selected_film_id)
    if not film:
        st.error("Film not found.")
        st.stop()
    repo = ReportRepository(session)
    report = build_film_report(repo, film)
    market_eg_price: float | None = repo.egypt_avg_ticket_price()

    fpf = session.get(FilmPerformanceFeatures, selected_film_id)
    fpf_data = {
        "eg_total_admissions": float(fpf.eg_total_admissions) if (fpf and fpf.eg_total_admissions is not None) else 0.0,
        "sa_total_admissions": float(fpf.sa_total_admissions) if (fpf and fpf.sa_total_admissions is not None) else 0.0,
        "ae_total_admissions": float(fpf.ae_total_admissions) if (fpf and fpf.ae_total_admissions is not None) else 0.0,
        "eg_stability": float(fpf.eg_stability) if (fpf and fpf.eg_stability is not None) else None,
        "sa_stability": float(fpf.sa_stability) if (fpf and fpf.sa_stability is not None) else None,
        "ae_stability": float(fpf.ae_stability) if (fpf and fpf.ae_stability is not None) else None,
    }

    ratings_rows = session.execute(
        select(RatingsMetric)
        .where(RatingsMetric.film_id == selected_film_id)
        .order_by(RatingsMetric.metric_date.desc(), RatingsMetric.created_at.desc())
    ).scalars().all()
    ratings_data = [
        {
            "source_name": (rr.source_name or "").strip().lower(),
            "rating_value": float(rr.rating_value) if rr.rating_value is not None else None,
            "vote_count": int(rr.vote_count) if rr.vote_count is not None else None,
        }
        for rr in ratings_rows
    ]

reconciled = dedupe_bom_weekend_rows(report.reconciled or [])
title = report.film.get("canonical_title", "")
year = report.film.get("release_year")
title_ar = report.film.get("canonical_title_ar")

# ── header ─────────────────────────────────────────────────────────────────────
st.markdown(f"## {title}" + (f" ({year})" if year else ""))
if title_ar:
    st.caption(title_ar)

ratings_line = format_ratings_line(report.ratings or [])
if ratings_line:
    st.caption(ratings_line)

section_divider()

# ── Filmyard data (Egypt-specific: daily + cumulative) ────────────────────────
filmyard_raw = report.raw_sections.get("Filmyard", [])
filmyard_daily_rows = [
    r for r in filmyard_raw
    if (r.get("record_granularity") or "").lower() in ("day", "daily")
    and float(r.get("period_gross_local") or 0) > 0
]
# Weekly native — only non-zero ones
filmyard_weekly_native = [
    r for r in filmyard_raw
    if (r.get("record_granularity") or "").lower() in ("week", "weekly")
    and float(r.get("period_gross_local") or 0) > 0
]
# Aggregate daily → weekly as fallback
filmyard_weekly_agg = _aggregate_filmyard_daily_to_weekly(filmyard_raw) if filmyard_raw else []

# Market-wide Egypt ticket price (from DB), fall back to film-specific
fy_price_source = filmyard_daily_rows  # daily rows have best gross/tickets ratio
film_eg_price: float | None = _calc_egypt_ticket_price(fy_price_source)
egypt_ticket_price: float | None = market_eg_price or film_eg_price

prices = load_prices()
settings = get_settings()


def _ticket_price(code: str) -> float | None:
    if code.upper() == "EG" and egypt_ticket_price:
        return egypt_ticket_price
    return prices.get(code.upper())


# Show Filmyard Egypt section if we have data
if filmyard_daily_rows:
    section_header("Filmyard — Egypt Live Data")
    st.caption("Direct from Filmyard box office tracker. Updates daily.")

    # Latest day's snapshot
    latest_daily = sorted(filmyard_daily_rows, key=lambda r: period_key_sort_ordinal(str(r.get("period_key") or "")), reverse=True)
    if latest_daily:
        ld = latest_daily[0]
        # Find cumulative total from any row that has it
        cumulative = next(
            (float(r["cumulative_gross_local"]) for r in sorted(filmyard_daily_rows, key=lambda r: period_key_sort_ordinal(str(r.get("period_key") or "")), reverse=True)
             if r.get("cumulative_gross_local") and float(r["cumulative_gross_local"]) > 0),
            None,
        )

        col1, col2, col3 = st.columns(3)
        with col1:
            daily_g = float(ld.get("period_gross_local") or 0)
            daily_adm = ld.get("admissions_actual")
            st.metric(
                f"Today ({format_period(str(ld.get('period_key') or ''))})",
                format_gross(daily_g, "EGP"),
                help="Today's gross from Filmyard",
            )
            if daily_adm is not None and float(daily_adm) > 0:
                st.caption(f"🎟 {int(float(daily_adm)):,} tickets today")

        with col2:
            # Best weekly source — native weekly if available, else daily-agg
            fy_weekly = filmyard_weekly_native if filmyard_weekly_native else filmyard_weekly_agg
            latest_wk = sorted(fy_weekly, key=lambda r: period_key_sort_ordinal(str(r.get("period_key") or "")), reverse=True)
            if latest_wk:
                lw = latest_wk[0]
                wk_g = float(lw.get("period_gross_local") or 0)
                wk_adm = lw.get("admissions_actual")
                wk_label = format_period(str(lw.get("period_key") or ""))
                src_note = "weekly" if filmyard_weekly_native else "daily aggregated"
                st.metric(
                    f"This Week ({wk_label})",
                    format_gross(wk_g, "EGP"),
                    help=f"Weekly gross ({src_note})",
                )
                if wk_adm is not None and float(wk_adm) > 0:
                    st.caption(f"🎟 {int(float(wk_adm)):,} tickets this week")
            else:
                st.metric("This Week", "—")

        with col3:
            if cumulative and cumulative > 0:
                st.metric(
                    "Cumulative Total",
                    format_gross(cumulative, "EGP"),
                    help="Running total from Filmyard",
                )
                if egypt_ticket_price and egypt_ticket_price > 0:
                    total_adm = int(cumulative / egypt_ticket_price)
                    st.caption(f"🎟 ~{total_adm:,} total tickets (est.)")
            else:
                st.metric("Cumulative Total", "—")

    st.markdown("")

# ── market filter ─────────────────────────────────────────────────────────────
# Filter out zero-gross and empty/invalid period_key reconciled rows
title_rows_all = [
    r for r in reconciled
    if r.get("semantics") == "title_period_gross"
    and r.get("period_key") not in ("lifetime", "", None)
    and float(r.get("period_gross_local") or 0) > 0  # hide zero-gross records
]
all_codes = sorted({(r.get("country_code") or "") for r in title_rows_all if r.get("country_code")})
all_market_labels = [market_name(c) for c in all_codes]
code_to_label = {c: market_name(c) for c in all_codes}
label_to_code = {v: k for k, v in code_to_label.items()}

if all_codes:
    selected_labels = st.multiselect("Markets to display", all_market_labels, default=all_market_labels)
    selected_codes = {label_to_code[l] for l in selected_labels}
else:
    selected_codes = set()

title_rows = [r for r in title_rows_all if (r.get("country_code") or "") in selected_codes]
totals_with_adm: list[dict] = []
display_rows: list[dict] = []
chart_images_for_export: dict[str, bytes] = {}

# ── market totals with estimated admissions ───────────────────────────────────
if title_rows:
    section_header("Box Office by Market")
    totals = _market_totals(title_rows)

    totals_with_adm = []
    for t_row in totals:
        mkt_label = t_row["Market"]
        code = label_to_code.get(mkt_label, "")
        mkt_gross_rows = [r for r in title_rows if market_name(r.get("country_code") or "") == mkt_label]
        total_gross_local = sum_gross_in_ticket_currency(mkt_gross_rows, code)
        total_actual_adm = sum(float(r.get("admissions_actual") or 0) for r in mkt_gross_rows if r.get("admissions_actual") is not None)
        has_actual = any(r.get("admissions_actual") is not None for r in mkt_gross_rows)

        if has_actual and total_actual_adm > 0:
            adm_str = f"{int(total_actual_adm):,} (actual)"
        elif total_gross_local > 0:
            price_override = egypt_ticket_price if code.upper() == "EG" else None
            gcur = ticket_currency_code(code)
            est = estimate_admissions(total_gross_local, code, price_override, gross_currency=gcur)
            adm_str = f"~{est:,} (est.)" if est else "-"
        else:
            adm_str = "-"

        totals_with_adm.append({**t_row, "Est. Admissions": adm_str})

    st.dataframe(totals_with_adm, use_container_width=True, hide_index=True)

    # Ticket price basis footnote
    price_notes = []
    for code in sorted(selected_codes):
        if code in prices or (code.upper() == "EG" and egypt_ticket_price):
            basis = price_basis_label(code, egypt_ticket_price if code.upper() == "EG" else None)
            price_notes.append(f"{market_name(code)}: {basis}")
    if price_notes:
        st.caption("Ticket price basis — " + " · ".join(price_notes))

    # ── weekly trend charts ───────────────────────────────────────────────────
    section_header("Week-by-Week Trend")
    chart_mode = st.radio(
        "Charts show",
        ["Gross", "Admissions"],
        horizontal=True,
        key="film_report_chart_metric",
        help="Gross uses period revenue. Admissions use actual ticket counts when present; otherwise an estimate from gross using ticket prices (see caption under the market totals table).",
    )
    metric = "admissions" if chart_mode == "Admissions" else "gross"
    if metric == "admissions":
        st.caption(
            "Admissions charts use **actual** counts when the source provides them; otherwise **~estimated** from gross "
            "and the ticket price basis shown above."
        )

    for code in sorted(selected_codes):
        currency = next((r.get("currency") or "" for r in title_rows if (r.get("country_code") or "").upper() == code), "")
        is_eg = code.upper() == "EG"
        y_title = (
            (f"Gross ({currency})" if currency else "Gross")
            if metric == "gross"
            else "Admissions"
        )

        # Egypt: elCinema + Filmyard on same chart
        if is_eg:
            fy_weekly = filmyard_weekly_native if filmyard_weekly_native else filmyard_weekly_agg
            elcinema_map: dict[str, float] = {}
            filmyard_map: dict[str, float] = {}
            ordinal_map: dict[str, int] = {}

            for r in title_rows:
                if (r.get("country_code") or "").upper() != "EG":
                    continue
                pk = str(r.get("period_key") or "")
                lbl = format_period_row(r)
                v = _row_value_for_chart(r, "EG", metric, egypt_ticket_price)
                if v is None or v <= 0:
                    continue
                elcinema_map[lbl] = v
                ordinal_map[lbl] = period_key_sort_ordinal(pk)

            for r in fy_weekly:
                pk = str(r.get("period_key") or "")
                lbl = format_period(pk)
                v = _row_value_for_chart(r, "EG", metric, egypt_ticket_price)
                if v is None or v <= 0:
                    continue
                filmyard_map[lbl] = v
                ordinal_map.setdefault(lbl, period_key_sort_ordinal(pk))

            all_labels = sorted(
                set(elcinema_map) | set(filmyard_map),
                key=lambda lbl: ordinal_map.get(lbl, 999999),
            )

            curr_lbl = f" ({currency})" if currency else " (EGP)"
            if not all_labels:
                st.markdown(f"**Egypt{curr_lbl}**")
                st.caption("No week-by-week data for this metric (try Gross, or check ticket prices for admissions).")
            else:
                combined_rows = []
                fy_col_name = "Filmyard (weekly)" if filmyard_weekly_native else "Filmyard (daily agg.)"
                for lbl in all_labels:
                    row: dict = {"label": lbl, "ordinal": ordinal_map.get(lbl, 999999)}
                    if lbl in elcinema_map:
                        row["elCinema"] = elcinema_map[lbl]
                    if lbl in filmyard_map:
                        row[fy_col_name] = filmyard_map[lbl]
                    combined_rows.append(row)
                st.markdown(f"**Egypt{curr_lbl}**")
                combined_df = pd.DataFrame(combined_rows)
                long_df = combined_df.melt(
                    id_vars=["label", "ordinal"],
                    var_name="source",
                    value_name="value",
                ).dropna(subset=["value"])
                if long_df.empty:
                    st.caption("No data to plot for this metric (try Gross, or check ticket prices for estimates).")
                else:
                    chart = _sorted_label_chart(
                        long_df,
                        "value",
                        color_col="source",
                        height=220,
                        y_title=y_title,
                    )
                    st.altair_chart(chart, use_container_width=True)
                    chart_labels = all_labels
                    series_vals: dict[str, list[float]] = {}
                    for src_name in long_df["source"].dropna().unique().tolist():
                        src_map = {
                            row["label"]: float(row["value"])
                            for _, row in long_df[long_df["source"] == src_name][["label", "value"]].iterrows()
                        }
                        series_vals[src_name] = [src_map.get(lbl, float("nan")) for lbl in chart_labels]
                    png = _build_market_chart_png("Egypt Trend", chart_labels, series_vals, y_title)
                    if png:
                        chart_images_for_export["Egypt"] = png
        else:
            chart_df = _chart_df(title_rows, code, metric, egypt_ticket_price)
            if chart_df is None or chart_df.empty:
                st.markdown(f"**{market_name(code)}**" + (f"  ({currency})" if currency else ""))
                st.caption("No data to plot for this metric (try Gross, or check ticket prices for admissions estimates).")
                continue
            label = f"{market_name(code)}" + (f"  ({currency})" if currency else "")
            st.markdown(f"**{label}**")
            chart = _sorted_label_chart(chart_df, "value", height=220, y_title=y_title)
            st.altair_chart(chart, use_container_width=True)
            chart_labels = chart_df["label"].tolist()
            chart_vals = chart_df["value"].astype(float).tolist()
            png = _build_market_chart_png(label.strip(), chart_labels, {"Series": chart_vals}, y_title)
            if png:
                chart_images_for_export[market_name(code)] = png

    # ── Egypt ticket price insight ────────────────────────────────────────────
    if egypt_ticket_price and "EG" in selected_codes:
        section_divider()
        section_header("Egypt Ticket Price (Filmyard-derived)")
        eg_col1, eg_col2 = st.columns(2)
        with eg_col1:
            st.metric(
                "Avg. Ticket Price — Egypt",
                f"{egypt_ticket_price:,.0f} EGP",
                help="Market-wide median of gross ÷ admissions from all Filmyard EG records.",
            )
        with eg_col2:
            eg_rows = [r for r in title_rows if (r.get("country_code") or "").upper() == "EG"]
            eg_total = sum_gross_in_ticket_currency(eg_rows, "EG")
            if eg_total > 0:
                est = estimate_admissions(eg_total, "EG", egypt_ticket_price, gross_currency=ticket_currency_code("EG"))
                if est:
                    st.metric("Est. Total Admissions — Egypt", f"{est:,}")

    # ── detailed weekly table ─────────────────────────────────────────────────
    with st.expander("Detailed data by period", expanded=False):
        detail_rows = []

        for r in sorted(
            title_rows,
            key=lambda x: (
                x.get("country_code") or "",
                period_key_sort_ordinal(str(x.get("period_key") or "")),
            ),
        ):
            code = r.get("country_code") or ""
            gross = float(r.get("period_gross_local") or 0)
            curr = r.get("currency") or ""
            adm_actual = r.get("admissions_actual")
            src = source_label(r.get("source") or "-", r.get("granularity") or "")
            period_key = str(r.get("period_key") or "")
            period_ord = period_key_sort_ordinal(period_key)

            if adm_actual is not None:
                adm_str = f"{int(float(adm_actual)):,} (actual)"
            elif gross > 0:
                price_override = egypt_ticket_price if code.upper() == "EG" else None
                gcur = (curr or "").strip() or None
                est = estimate_admissions(gross, code, price_override, gross_currency=gcur)
                adm_str = f"~{est:,} (est.)" if est else "-"
            else:
                adm_str = "-"

            detail_rows.append({
                "Market": market_name(code),
                "Period": format_period_row(r),
                "Gross": format_gross(gross, curr),
                "Admissions": adm_str,
                "Source": src,
                "_market_sort": code,
                "_period_sort": period_ord,
                "_source_sort": src,
                "_is_cumulative": 0,
            })

        # Add Filmyard daily rows explicitly for Egypt (cumulative rows only if not already present)
        if filmyard_daily_rows and "EG" in selected_codes:
            existing_keys = {(r["Market"], r["Period"], r["Source"]) for r in detail_rows}
            for r in sorted(filmyard_daily_rows, key=lambda x: period_key_sort_ordinal(str(x.get("period_key") or ""))):
                gross = float(r.get("period_gross_local") or 0)
                adm = r.get("admissions_actual")
                cumul = r.get("cumulative_gross_local")
                period_lbl = format_period(str(r.get("period_key") or ""))

                if adm is not None and float(adm) > 0:
                    adm_str = f"{int(float(adm)):,} (actual)"
                elif gross > 0:
                    est = estimate_admissions(gross, "EG", egypt_ticket_price)
                    adm_str = f"~{est:,} (est.)" if est else "-"
                else:
                    adm_str = "-"

                daily_key = ("Egypt", period_lbl, "Filmyard (daily)")
                if daily_key not in existing_keys:
                    detail_rows.append({
                        "Market": "Egypt",
                        "Period": period_lbl,
                        "Gross": format_gross(gross, "EGP"),
                        "Admissions": adm_str,
                        "Source": "Filmyard (daily)",
                        "_market_sort": "EG",
                        "_period_sort": period_key_sort_ordinal(str(r.get("period_key") or "")),
                        "_source_sort": "Filmyard (daily)",
                        "_is_cumulative": 0,
                    })
                    existing_keys.add(daily_key)
                # Cumulative total row — always add (never in reconciled)
                if cumul and float(cumul) > 0:
                    cum_adm = int(float(cumul) / egypt_ticket_price) if egypt_ticket_price else None
                    detail_rows.append({
                        "Market": "Egypt",
                        "Period": f"{period_lbl} cumul.",
                        "Gross": format_gross(float(cumul), "EGP"),
                        "Admissions": f"~{cum_adm:,} (est.)" if cum_adm else "-",
                        "Source": "Filmyard (cumulative)",
                        "_market_sort": "EG",
                        "_period_sort": period_key_sort_ordinal(str(r.get("period_key") or "")),
                        "_source_sort": "Filmyard (cumulative)",
                        "_is_cumulative": 1,
                    })

        if detail_rows:
            sorted_rows = sorted(
                detail_rows,
                key=lambda x: (
                    x.get("_market_sort") or "",
                    int(x.get("_period_sort") or 0),
                    int(x.get("_is_cumulative") or 0),
                    x.get("_source_sort") or "",
                ),
            )
            display_rows = [
                {
                    "Market": r["Market"],
                    "Period": r["Period"],
                    "Gross": r["Gross"],
                    "Admissions": r["Admissions"],
                    "Source": r["Source"],
                }
                for r in sorted_rows
            ]
            st.dataframe(
                display_rows,
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.caption("No detailed data available.")
else:
    st.info("No box office data available for this film yet. Use Film Search to fetch it.")

# ── marketing inputs & targets ────────────────────────────────────────────────
mkt_inputs = report.marketing_inputs or []
targets = report.outcome_targets or []

if mkt_inputs or targets:
    section_divider()
    section_header("Campaign Budget & Targets")
    col1, col2 = st.columns(2)
    with col1:
        if mkt_inputs:
            st.markdown("**Marketing Spend**")
            st.dataframe(
                [{"Market": market_name(m["market_code"]),
                  "Spend": format_gross(float(m["spend_local"] or 0), m["spend_currency"] or "")}
                 for m in mkt_inputs],
                use_container_width=True, hide_index=True,
            )
    with col2:
        if targets:
            st.markdown("**Admissions Target**")
            st.dataframe(
                [{"Market": market_name(t["market_code"]),
                  "Target": f"{float(t['target_value'] or 0):,.0f} {t['target_unit'] or ''}".strip()}
                 for t in targets],
                use_container_width=True, hide_index=True,
            )

# ── investment analysis ───────────────────────────────────────────────────────
section_divider()
section_header("Investment Analysis")

eg_total_adm = float(fpf_data.get("eg_total_admissions") or 0.0)
sa_total_adm = float(fpf_data.get("sa_total_admissions") or 0.0)
ae_total_adm = float(fpf_data.get("ae_total_admissions") or 0.0)

stabilities = []
if fpf_data.get("eg_stability") is not None:
    stabilities.append(float(fpf_data["eg_stability"]))
if fpf_data.get("sa_stability") is not None:
    stabilities.append(float(fpf_data["sa_stability"]))
if fpf_data.get("ae_stability") is not None:
    stabilities.append(float(fpf_data["ae_stability"]))
avg_stability = (sum(stabilities) / len(stabilities)) if stabilities else 0.0

imdb_rating = None
elcinema_rating = None
letterboxd_rating = None
letterboxd_votes = None
for rr in ratings_data:
    src = rr["source_name"]
    if imdb_rating is None and src == "imdb" and rr["rating_value"] is not None:
        imdb_rating = float(rr["rating_value"])
    if elcinema_rating is None and src == "elcinema" and rr["rating_value"] is not None:
        elcinema_rating = float(rr["rating_value"])
    if letterboxd_rating is None and src == "letterboxd" and rr["rating_value"] is not None:
        letterboxd_rating = float(rr["rating_value"])
    if letterboxd_votes is None and src == "letterboxd" and rr["vote_count"] is not None:
        letterboxd_votes = int(rr["vote_count"])
    if imdb_rating is not None and elcinema_rating is not None and letterboxd_rating is not None and letterboxd_votes is not None:
        break

# Film features — displayed as KPI cards
st.caption("Film features used for prediction:")
col_a, col_b, col_c = st.columns(3)
with col_a:
    kpi_card("Egypt Admissions", f"{eg_total_adm:,.0f}")
with col_b:
    kpi_card("Saudi Admissions", f"{sa_total_adm:,.0f}")
with col_c:
    kpi_card("UAE Admissions", f"{ae_total_adm:,.0f}")

st.markdown("")
col_d, col_e, col_f, col_g = st.columns(4)
with col_d:
    kpi_card("Avg Stability", f"{avg_stability:.3f}")
with col_e:
    kpi_card("IMDb", f"{(imdb_rating if imdb_rating is not None else 0):.1f}/10")
with col_f:
    kpi_card("elCinema", f"{(elcinema_rating if elcinema_rating is not None else 0):.1f}/10")
with col_g:
    kpi_card("Letterboxd", f"{(letterboxd_rating if letterboxd_rating is not None else 0):.2f}/5",
             f"{int(letterboxd_votes or 0):,} votes")

st.markdown("")
marketing_budget = st.number_input(
    "Marketing Budget",
    min_value=0.0,
    value=0.0,
    step=10_000.0,
    help="Budget sent to the prediction API as `marketing_spend`.",
)

if st.button("Run Investment Analysis", type="primary", use_container_width=True):
    payload = {
        "eg_adm": eg_total_adm,
        "sa_adm": sa_total_adm,
        "ae_adm": ae_total_adm,
        "imdb": float(imdb_rating or 0.0),
        "elcinema": float(elcinema_rating or 0.0),
        "letterboxd_rating": float(letterboxd_rating or 0.0),
        "letterboxd_votes": int(letterboxd_votes or 0),
        "stability": float(avg_stability or 0.0),
        "marketing_spend": float(marketing_budget),
    }
    try:
        if settings.prediction_mode.lower() == "api":
            result = predict_via_api(payload, "predict")
        else:
            result = run_investment_analysis(
                InferenceInput(
                    eg_adm=float(payload["eg_adm"]),
                    sa_adm=float(payload["sa_adm"]),
                    ae_adm=float(payload["ae_adm"]),
                    imdb=float(payload["imdb"]),
                    elcinema=float(payload["elcinema"]),
                    letterboxd_rating=float(payload["letterboxd_rating"]),
                    letterboxd_votes=int(payload["letterboxd_votes"]),
                    stability=float(payload["stability"]),
                    marketing_spend=float(payload["marketing_spend"]),
                )
            )
    except Exception as exc:  # noqa: BLE001
        st.error(f"Prediction error: {exc}")
        result = None

    if result:
        pred_fw = float(result.get("predicted_first_watch") or 0.0)
        roi_val = float(result.get("roi") or 0.0)
        rev_val = float(result.get("revenue") or 0.0)
        prof_val = float(result.get("profit") or 0.0)
        decision = str(result.get("decision") or "-")

        section_header("Analysis Result")
        r1, r2, r3, r4 = st.columns(4)
        with r1:
            kpi_card("Predicted First Watch", f"{pred_fw:,.0f}")
        with r2:
            kpi_card("ROI", f"{roi_val:.2f}")
        with r3:
            kpi_card("Revenue", f"{rev_val:,.0f}")
        with r4:
            kpi_card("Profit", f"{prof_val:,.0f}")

        st.markdown("")
        decision_badge(decision)
        st.session_state[f"investment_result_{selected_film_id}"] = {
            "Predicted First Watch": f"{pred_fw:,.0f}",
            "ROI": f"{roi_val:.2f}",
            "Revenue": f"{rev_val:,.0f}",
            "Profit": f"{prof_val:,.0f}",
            "Decision": decision,
            "Input Marketing Budget": f"{marketing_budget:,.0f}",
        }

        with session_scope() as s:
            s.add(
                FilmInvestmentAnalysis(
                    film_id=selected_film_id,
                    predicted_first_watch=Decimal(str(pred_fw)),
                    suggested_marketing_spend=Decimal(str(marketing_budget)),
                    roi=Decimal(str(roi_val)),
                    estimated_revenue=Decimal(str(rev_val)),
                    estimated_profit=Decimal(str(prof_val)),
                    decision=decision,
                )
            )
        st.caption("Saved to film_investment_analysis.")

section_divider()
section_header("Target First Watch")
target_first_watch = st.number_input(
    "Target First Watch",
    min_value=0.0,
    value=0.0,
    step=1_000.0,
    help="Enter your target first-watch result to estimate required marketing budget.",
)

if st.button("Calculate Required Marketing Budget", type="secondary", use_container_width=True):
    payload_target = {
        "eg_adm": eg_total_adm,
        "sa_adm": sa_total_adm,
        "ae_adm": ae_total_adm,
        "imdb": float(imdb_rating or 0.0),
        "elcinema": float(elcinema_rating or 0.0),
        "letterboxd_rating": float(letterboxd_rating or 0.0),
        "letterboxd_votes": int(letterboxd_votes or 0),
        "stability": float(avg_stability or 0.0),
        "target_first_watch": float(target_first_watch),
    }
    try:
        if settings.prediction_mode.lower() == "api":
            result_target = predict_via_api(payload_target, "suggest-spend")
        else:
            result_target = suggest_spend_from_target(
                target_first_watch=float(payload_target["target_first_watch"]),
                eg_adm=float(payload_target["eg_adm"]),
                sa_adm=float(payload_target["sa_adm"]),
                ae_adm=float(payload_target["ae_adm"]),
                imdb=float(payload_target["imdb"]),
                elcinema=float(payload_target["elcinema"]),
                letterboxd_rating=float(payload_target["letterboxd_rating"]),
                letterboxd_votes=int(payload_target["letterboxd_votes"]),
                stability=float(payload_target["stability"]),
            )
    except Exception as exc:  # noqa: BLE001
        st.error(f"Suggest-spend error: {exc}")
        result_target = None

    if result_target:
        if result_target.get("error"):
            st.error(str(result_target["error"]))
        else:
            required_budget = float(result_target.get("suggested_marketing_spend") or 0.0)
            section_header("Required Budget")
            kpi_card("Required Marketing Budget", f"{required_budget:,.0f}")
            st.session_state[f"target_watch_result_{selected_film_id}"] = {
                "Target First Watch": f"{target_first_watch:,.0f}",
                "Required Marketing Budget": f"{required_budget:,.0f}",
            }

section_divider()
section_header("Download Summary")
st.caption("Export a clean, shareable report. Investment and target sections are included only if you ran them.")

inv_state = st.session_state.get(f"investment_result_{selected_film_id}")
target_state = st.session_state.get(f"target_watch_result_{selected_film_id}")
summary_rows = [
    {"Item": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M")},
    {"Item": "Film", "Value": f"{title} ({year})" if year else title},
    {"Item": "Displayed Markets", "Value": ", ".join(sorted(selected_labels)) if all_codes else "None"},
    {"Item": "Ratings Snapshot", "Value": ratings_line or "Not available"},
    {"Item": "Investment Analysis Included", "Value": "Yes" if inv_state else "No"},
    {"Item": "Target First Watch Included", "Value": "Yes" if target_state else "No"},
]

marketing_rows = [
    {"Market": market_name(m["market_code"]), "Spend": format_gross(float(m["spend_local"] or 0), m["spend_currency"] or "")}
    for m in mkt_inputs
]
target_rows = [
    {"Market": market_name(t["market_code"]), "Target": f"{float(t['target_value'] or 0):,.0f} {t['target_unit'] or ''}".strip()}
    for t in targets
]
investment_rows = [inv_state] if inv_state else []
target_result_rows = [target_state] if target_state else []

excel_bytes = _build_excel_report(
    film_title=f"{title} ({year})" if year else title,
    summary_rows=summary_rows,
    totals_rows=totals_with_adm,
    detail_rows=display_rows,
    marketing_rows=marketing_rows,
    target_rows=target_rows,
    investment_rows=investment_rows,
    target_result_rows=target_result_rows,
    chart_images=chart_images_for_export,
)
pdf_bytes = _build_pdf_report(
    film_title=f"{title} ({year})" if year else title,
    summary_rows=summary_rows,
    totals_rows=totals_with_adm,
    investment_rows=investment_rows,
    target_result_rows=target_result_rows,
    chart_images=chart_images_for_export,
)

dcol1, dcol2 = st.columns(2)
with dcol1:
    st.download_button(
        "Download Excel Summary (.xlsx)",
        data=excel_bytes,
        file_name=f"{title.replace(' ', '_')}_film_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with dcol2:
    st.download_button(
        "Download PDF Summary (.pdf)",
        data=pdf_bytes,
        file_name=f"{title.replace(' ', '_')}_film_report.pdf",
        mime="application/pdf",
        use_container_width=True,
    )
